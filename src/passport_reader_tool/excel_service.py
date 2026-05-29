from __future__ import annotations

from datetime import date
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from passport_reader_tool.models import PassportRecord


EXCEL_HEADERS = [
    "STT",
    "Họ tên",
    "Ngày sinh",
    "Giới tính",
    "Số hộ chiếu",
    "Ngày hết hạn",
    "Ngày thêm",
]

SHEET_NAME = "Passport Data"
DATE_FORMAT = "dd/mm/yyyy"
ERROR_FILL = PatternFill(fill_type="solid", fgColor="FFC7CE")
HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")


class ExcelWorkbookService:
    def new_workbook(self) -> Workbook:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = SHEET_NAME
        self._write_header(worksheet)
        self._set_column_widths(worksheet)
        return workbook

    def load(self, path: str | Path) -> tuple[Workbook, list[PassportRecord]]:
        workbook = load_workbook(path)
        worksheet = self._get_or_create_sheet(workbook)
        self._ensure_header(worksheet)
        records = list(self._read_records(worksheet))
        return workbook, records

    def save(self, workbook: Workbook, path: str | Path, records: Iterable[PassportRecord]) -> None:
        worksheet = self._get_or_create_sheet(workbook)
        self._clear_sheet(worksheet)
        self._write_header(worksheet)
        for record in records:
            self._append_record(worksheet, record)
        self._set_column_widths(worksheet)
        workbook.save(path)

    def append_to_workbook(self, workbook: Workbook, records: Iterable[PassportRecord]) -> None:
        worksheet = self._get_or_create_sheet(workbook)
        self._ensure_header(worksheet)
        for record in records:
            self._append_record(worksheet, record)

    def _get_or_create_sheet(self, workbook: Workbook) -> Worksheet:
        if SHEET_NAME in workbook.sheetnames:
            return workbook[SHEET_NAME]
        worksheet = workbook.active
        worksheet.title = SHEET_NAME
        return worksheet

    def _ensure_header(self, worksheet: Worksheet) -> None:
        current = [worksheet.cell(row=1, column=i + 1).value for i in range(len(EXCEL_HEADERS))]
        if current != EXCEL_HEADERS:
            self._write_header(worksheet)
            self._set_column_widths(worksheet)

    def _write_header(self, worksheet: Worksheet) -> None:
        for index, header in enumerate(EXCEL_HEADERS, start=1):
            cell = worksheet.cell(row=1, column=index, value=header)
            cell.font = Font(bold=True)
            cell.fill = HEADER_FILL
        worksheet.freeze_panes = "A2"

    def _append_record(self, worksheet: Worksheet, record: PassportRecord) -> None:
        row = [
            record.row_number,
            record.full_name,
            record.date_of_birth,
            record.sex,
            record.passport_number,
            record.expiry_date,
            record.added_date,
        ]
        worksheet.append(row)
        row_index = worksheet.max_row
        for column_index in (3, 6, 7):
            worksheet.cell(row=row_index, column=column_index).number_format = DATE_FORMAT
        if record.is_error:
            for column_index in range(1, len(EXCEL_HEADERS) + 1):
                worksheet.cell(row=row_index, column=column_index).fill = ERROR_FILL

    def _read_records(self, worksheet: Worksheet) -> Iterable[PassportRecord]:
        for row_index in range(2, worksheet.max_row + 1):
            values = [worksheet.cell(row=row_index, column=i).value for i in range(1, len(EXCEL_HEADERS) + 1)]
            if not any(values):
                continue
            yield PassportRecord(
                row_number=int(values[0] or row_index - 1),
                full_name=str(values[1] or ""),
                date_of_birth=self._date_or_none(values[2]),
                sex=str(values[3] or ""),
                passport_number=str(values[4] or ""),
                expiry_date=self._date_or_none(values[5]),
                added_date=self._date_or_none(values[6]),
            )

    def _clear_sheet(self, worksheet: Worksheet) -> None:
        if worksheet.max_row:
            worksheet.delete_rows(1, worksheet.max_row)

    def _set_column_widths(self, worksheet: Worksheet) -> None:
        widths = [8, 28, 14, 12, 18, 16, 14]
        for index, width in enumerate(widths, start=1):
            worksheet.column_dimensions[worksheet.cell(row=1, column=index).column_letter].width = width

    def _date_or_none(self, value: object) -> date | None:
        if isinstance(value, date):
            return value
        return None
