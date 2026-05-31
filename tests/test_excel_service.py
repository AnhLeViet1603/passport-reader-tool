from datetime import date

from openpyxl import load_workbook

from passport_reader_tool.excel_service import EXCEL_HEADERS, OLD_EXCEL_HEADERS, ExcelWorkbookService
from passport_reader_tool.models import PassportRecord


def test_save_workbook_writes_headers_dates_and_error_fill(tmp_path):
    output = tmp_path / "passports.xlsx"
    service = ExcelWorkbookService()
    workbook = service.new_workbook()
    records = [
        PassportRecord(
            row_number=1,
            full_name="NGUYEN VAN A",
            date_of_birth=date(1990, 1, 2),
            sex="M",
            passport_number="B1234567",
            issue_date=date(2020, 2, 3),
            expiry_date=date(2030, 3, 4),
            added_date=date(2026, 5, 29),
            name_mismatch=True,
        ),
        PassportRecord(row_number=2, added_date=date(2026, 5, 29), status="error", error_message="OCR fail"),
    ]

    service.save(workbook, output, records)

    saved = load_workbook(output)
    sheet = saved["Passport Data"]
    assert [sheet.cell(row=1, column=i).value for i in range(1, len(EXCEL_HEADERS) + 1)] == EXCEL_HEADERS
    assert sheet.cell(row=2, column=2).value == "NGUYEN VAN A"
    assert sheet.cell(row=2, column=3).number_format == "dd/mm/yyyy"
    assert sheet.cell(row=2, column=6).value.date() == date(2020, 2, 3)
    assert sheet.cell(row=2, column=2).fill.fgColor.rgb == "00FFEB9C"
    assert sheet.cell(row=3, column=1).fill.fgColor.rgb == "00FFC7CE"


def test_load_workbook_reads_records(tmp_path):
    output = tmp_path / "passports.xlsx"
    service = ExcelWorkbookService()
    workbook = service.new_workbook()
    service.save(
        workbook,
        output,
        [
            PassportRecord(
                row_number=1,
                full_name="TRAN THI B",
                date_of_birth=date(1988, 6, 7),
                sex="F",
                passport_number="C7654321",
                issue_date=date(2019, 7, 8),
                expiry_date=date(2029, 8, 9),
                added_date=date(2026, 5, 29),
            )
        ],
    )

    _, records = service.load(output)

    assert len(records) == 1
    assert records[0].full_name == "TRAN THI B"
    assert records[0].date_of_birth == date(1988, 6, 7)
    assert records[0].issue_date == date(2019, 7, 8)


def test_load_workbook_migrates_old_header_by_inserting_issue_date(tmp_path):
    output = tmp_path / "old-passports.xlsx"
    service = ExcelWorkbookService()
    workbook = service.new_workbook()
    sheet = workbook["Passport Data"]
    sheet.delete_cols(6)
    for index, header in enumerate(OLD_EXCEL_HEADERS, start=1):
        sheet.cell(row=1, column=index, value=header)
    sheet.append([1, "LE VAN C", date(1992, 1, 2), "M", "D1234567", date(2032, 3, 4), date(2026, 5, 31)])
    workbook.save(output)

    loaded_workbook, records = service.load(output)

    assert records[0].issue_date is None
    assert records[0].expiry_date == date(2032, 3, 4)
    sheet = loaded_workbook["Passport Data"]
    assert [sheet.cell(row=1, column=i).value for i in range(1, len(EXCEL_HEADERS) + 1)] == EXCEL_HEADERS
